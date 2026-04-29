"""
Full Pipeline Test: Tavily Search → Firecrawl Listing → Firecrawl Vendor Profiles → Excel

Flow:
  1. Tavily search → seed expo URLs
  2. Filter kandidat listing page terbaik
  3. Firecrawl scrape listing + links → discover vendor profile URLs
  4. Firecrawl scrape tiap vendor profile → extract data terstruktur
  5. Export Excel (per-step visibility)

Usage:
    python test_pipeline_tavily_firecrawl.py

Keys dari .env:
    TAVILY_API_KEY=tvly-xxxxx
    FIRECRAWL_API_KEY=fc-xxxxx

Credit usage (Firecrawl):
    ~3 listing pages × 1 + ~30 vendor profiles × 1 = ~33 credits
"""

import sys
import io
import os
import re
import time
import json
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from urllib.parse import urlparse, urljoin

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

import httpx
from rich.console import Console
from rich.table import Table
from rich.panel import Panel
from rich.progress import (
    Progress, SpinnerColumn, TextColumn, TimeElapsedColumn,
)
from rich import box
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

# ──────────────────────────────────────────────────────────────────────────────
# .env loader
# ──────────────────────────────────────────────────────────────────────────────

def _load_env() -> None:
    env_path = Path(__file__).resolve().parent / ".env"
    if not env_path.exists():
        return
    with open(env_path, encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            key, _, val = line.partition("=")
            key = key.strip()
            val = val.strip().strip('"').strip("'")
            if key and key not in os.environ:
                os.environ[key] = val

_load_env()

TAVILY_API_KEY    = os.environ.get("TAVILY_API_KEY", "")
FIRECRAWL_API_KEY = os.environ.get("FIRECRAWL_API_KEY", "")

OUTPUT_DIR = Path("output")
console    = Console()

# ──────────────────────────────────────────────────────────────────────────────
# Config — ubah sesuai kebutuhan
# ──────────────────────────────────────────────────────────────────────────────

SEARCH_QUERY        = "defense aerospace exhibition exhibitors list 2025 2026"
TAVILY_MAX_RESULTS  = 10          # hasil per Tavily search
MAX_LISTING_PAGES   = 3           # berapa listing page yang di-scrape
MAX_PROFILES_PER_LISTING = 10     # berapa vendor profile per listing (hemat kredit)

VENDOR_URL_KEYWORDS = [
    "exhibitor", "vendor", "company", "booth", "brand", "participant",
    "sponsor", "supplier", "partner", "profile", "member", "detail",
    "directory", "listing", "katilimci", "firma", "expositor",
]

NAV_SKIP_KEYWORDS = [
    "login", "register", "signup", "contact", "about", "faq",
    "privacy", "terms", "cookie", "sitemap", "search", "tag",
    "news", "blog", "press", "media", "career", "jobs", "help",
    "cart", "checkout", "account", "logout",
]

LISTING_URL_SIGNALS = [
    "exhibitors", "vendors", "participants", "sponsors", "companies",
    "directory", "floor-plan", "floorplan", "booth", "listing",
]

# Firecrawl extract schema — field yang kita minta
VENDOR_EXTRACT_SCHEMA = {
    "type": "object",
    "properties": {
        "company_name": {"type": "string", "description": "Official company / vendor name"},
        "website":      {"type": "string", "description": "Official company website URL"},
        "email":        {"type": "string", "description": "Contact email address"},
        "phone":        {"type": "string", "description": "Phone number"},
        "country":      {"type": "string", "description": "Country of the company"},
        "city":         {"type": "string", "description": "City of the company"},
        "description":  {"type": "string", "description": "Short company description or product focus"},
        "booth_number": {"type": "string", "description": "Booth, stand, or hall number at the event"},
        "category":     {"type": "string", "description": "Industry or product category"},
        "linkedin":     {"type": "string", "description": "LinkedIn company page URL"},
    },
    "required": [],
}

# ──────────────────────────────────────────────────────────────────────────────
# Excel styles
# ──────────────────────────────────────────────────────────────────────────────

NAVY   = PatternFill("solid", fgColor="1F3864")
BLUE2  = PatternFill("solid", fgColor="2E75B6")
GREEN2 = PatternFill("solid", fgColor="375623")
ORANGE = PatternFill("solid", fgColor="843C0C")
STRIPE = PatternFill("solid", fgColor="DCE6F1")
STRIPE2= PatternFill("solid", fgColor="E2EFDA")
HDR_FONT  = Font(bold=True, color="FFFFFF", size=11)
BODY_FONT = Font(size=10)
CENTER    = Alignment(horizontal="center", vertical="center", wrap_text=True)
TOP_WRAP  = Alignment(vertical="top", wrap_text=True)


def _write_header(ws, columns: list[tuple[str, int]], fill=NAVY) -> None:
    for col, (label, width) in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col, value=label)
        cell.fill      = fill
        cell.font      = HDR_FONT
        cell.alignment = CENTER
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"

# ──────────────────────────────────────────────────────────────────────────────
# STEP 1 — Tavily Search
# ──────────────────────────────────────────────────────────────────────────────

def tavily_search(query: str, max_results: int = 10) -> list[dict]:
    if not TAVILY_API_KEY:
        return []
    body = {
        "api_key":      TAVILY_API_KEY,
        "query":        query,
        "max_results":  max_results,
        "search_depth": "basic",
        "include_answer":      False,
        "include_raw_content": False,
    }
    try:
        resp = httpx.post("https://api.tavily.com/search", json=body, timeout=30)
        if resp.status_code == 200:
            return resp.json().get("results", [])
    except Exception as e:
        console.print(f"[red]Tavily error: {e}[/red]")
    return []


def _score_as_listing(url: str, title: str = "", snippet: str = "") -> int:
    """Score seberapa besar kemungkinan URL ini adalah halaman listing vendor."""
    combined = (url + " " + title + " " + snippet).lower()
    score = 0
    for kw in LISTING_URL_SIGNALS:
        if kw in url.lower():
            score += 3
        elif kw in combined:
            score += 1
    return score


def step1_tavily_search(query: str) -> list[dict]:
    console.print(Panel.fit(
        f"[bold cyan]STEP 1 — Tavily Search[/bold cyan]\n[dim]{query}[/dim]",
        border_style="cyan",
    ))

    with Progress(SpinnerColumn(), TextColumn("{task.description}"), console=console) as p:
        t = p.add_task("Searching Tavily ...", total=None)
        results = tavily_search(query, TAVILY_MAX_RESULTS)
        p.remove_task(t)

    candidates = []
    for r in results:
        url     = r.get("url", "")
        title   = r.get("title", "")
        snippet = r.get("content", "")
        score   = _score_as_listing(url, title, snippet)
        candidates.append({
            "url":     url,
            "title":   title,
            "snippet": snippet[:200],
            "score":   r.get("score", 0.0),
            "listing_score": score,
        })

    # Sort by listing_score × tavily_score
    candidates.sort(key=lambda x: x["listing_score"] * x["score"], reverse=True)

    table = Table(box=box.SIMPLE_HEAVY, header_style="bold white on dark_blue", show_lines=True)
    table.add_column("#",              justify="center", width=4)
    table.add_column("Tavily Score",   justify="center", width=13)
    table.add_column("Listing Score",  justify="center", width=14)
    table.add_column("URL",            min_width=60, style="cyan")
    table.add_column("Title",          min_width=40)

    for i, c in enumerate(candidates, 1):
        table.add_row(
            str(i),
            f"{c['score']:.3f}",
            str(c["listing_score"]),
            c["url"][:60],
            c["title"][:40],
        )

    console.print(table)
    console.print(f"  [bold]Kandidat dari Tavily:[/bold] [green]{len(candidates)}[/green]\n")
    return candidates

# ──────────────────────────────────────────────────────────────────────────────
# STEP 2 — Firecrawl: Scrape Listing Page + Get Links
# ──────────────────────────────────────────────────────────────────────────────

def firecrawl_scrape(url: str, want_links: bool = False, want_extract: bool = False) -> dict:
    """Scrape satu URL via Firecrawl. Return dict dengan markdown, links, extract."""
    if not FIRECRAWL_API_KEY:
        return {"success": False, "error": "FIRECRAWL_API_KEY not set", "url": url}

    formats = ["markdown"]
    if want_links:
        formats.append("links")
    if want_extract:
        formats.append("extract")

    body: dict = {"url": url, "formats": formats}
    if want_extract:
        body["extract"] = {"schema": VENDOR_EXTRACT_SCHEMA}

    headers = {
        "Authorization": f"Bearer {FIRECRAWL_API_KEY}",
        "Content-Type":  "application/json",
    }
    t0 = time.perf_counter()
    try:
        resp    = httpx.post("https://api.firecrawl.dev/v1/scrape", headers=headers, json=body, timeout=60)
        elapsed = time.perf_counter() - t0

        if resp.status_code == 200:
            data  = resp.json()
            inner = data.get("data", {})
            meta  = inner.get("metadata", {})
            return {
                "url":      url,
                "success":  data.get("success", False),
                "markdown": inner.get("markdown", ""),
                "links":    inner.get("links", []),
                "extract":  inner.get("extract", {}) or {},
                "title":    meta.get("title", ""),
                "elapsed":  round(elapsed, 2),
                "error":    "",
            }
        else:
            return {
                "url": url, "success": False, "markdown": "", "links": [],
                "extract": {}, "title": "",
                "elapsed": round(elapsed, 2),
                "error": f"HTTP {resp.status_code}: {resp.text[:200]}",
            }
    except Exception as e:
        return {
            "url": url, "success": False, "markdown": "", "links": [],
            "extract": {}, "title": "",
            "elapsed": round(time.perf_counter() - t0, 2),
            "error": f"{type(e).__name__}: {e}",
        }


def _discover_vendor_urls_from_links(links: list[str], base_url: str, max_urls: int) -> list[str]:
    """
    Dari daftar links hasil Firecrawl, temukan URL yang kemungkinan vendor profile.
    Strategy: keyword score + pattern repetition bonus.
    """
    parsed_base = urlparse(base_url)
    base_domain = parsed_base.netloc
    source_path = base_url.rstrip("/")

    scores:          dict[str, int]        = {}
    pattern_groups:  dict[str, list[str]]  = defaultdict(list)

    for href in links:
        if not href or not href.startswith("http"):
            continue
        p = urlparse(href)
        if base_domain not in p.netloc:
            continue
        if href.rstrip("/") == source_path:
            continue
        # Skip long query strings (filters / search)
        if len(p.query) > 40:
            continue

        path_lower = p.path.lower()

        # Skip nav pages
        if any(kw in path_lower for kw in NAV_SKIP_KEYWORDS):
            continue

        # Keyword score
        score = 0
        for kw in VENDOR_URL_KEYWORDS:
            if kw in path_lower:
                score += 3

        # Group by first 2 path segments
        parts = [x for x in p.path.split("/") if x]
        if len(parts) >= 2:
            pat = "/" + "/".join(parts[:2])
        elif parts:
            pat = "/" + parts[0]
        else:
            continue  # skip root

        pattern_groups[pat].append(href)
        scores[href] = score

    # Bonus for repeated patterns (listing grid)
    for pat, hrefs in pattern_groups.items():
        if len(hrefs) >= 3:
            for h in hrefs:
                scores[h] = scores.get(h, 0) + 5

    # Sort by score desc, deduplicate
    ranked = sorted(scores.items(), key=lambda x: x[1], reverse=True)
    seen   = set()
    result = []
    for url, sc in ranked:
        if url not in seen and sc > 0:
            seen.add(url)
            result.append(url)
        if len(result) >= max_urls:
            break

    return result


def step2_scrape_listings(candidates: list[dict]) -> list[dict]:
    """
    Scrape listing pages dengan Firecrawl.
    Return list of {listing_url, vendor_urls, firecrawl_result}.
    """
    console.print(Panel.fit(
        f"[bold yellow]STEP 2 — Firecrawl: Scrape Listing Pages[/bold yellow]\n"
        f"[dim]{min(MAX_LISTING_PAGES, len(candidates))} halaman terbaik[/dim]",
        border_style="yellow",
    ))

    listing_results = []
    top_candidates  = candidates[:MAX_LISTING_PAGES]

    with Progress(SpinnerColumn(), TextColumn("{task.description:<65}"), TimeElapsedColumn(), console=console) as p:
        task = p.add_task("Scraping ...", total=len(top_candidates))

        for cand in top_candidates:
            url = cand["url"]
            p.update(task, description=f"[yellow]{url[:62]}[/yellow]")

            fc = firecrawl_scrape(url, want_links=True, want_extract=False)
            time.sleep(1.0)

            vendor_urls = []
            if fc["success"] and fc.get("links"):
                vendor_urls = _discover_vendor_urls_from_links(
                    fc["links"], url, MAX_PROFILES_PER_LISTING
                )

            listing_results.append({
                "listing_url":  url,
                "listing_title": cand.get("title", ""),
                "fc_title":     fc.get("title", ""),
                "markdown_len": len(fc.get("markdown", "")),
                "links_found":  len(fc.get("links", [])),
                "vendor_urls":  vendor_urls,
                "fc_success":   fc["success"],
                "fc_elapsed":   fc["elapsed"],
                "fc_error":     fc.get("error", ""),
            })

            p.advance(task)

    # Summary
    table = Table(box=box.SIMPLE_HEAVY, header_style="bold white on dark_orange3", show_lines=True)
    table.add_column("Listing URL",      min_width=55, style="yellow")
    table.add_column("Links",            justify="center", width=8)
    table.add_column("Vendor URLs",      justify="center", width=13)
    table.add_column("MD Len",           justify="center", width=9)
    table.add_column("Status",           justify="center", width=12)

    for lr in listing_results:
        ok     = lr["fc_success"]
        status = "[green]✅[/green]" if ok else f"[red]❌[/red]"
        table.add_row(
            lr["listing_url"][:55],
            str(lr["links_found"]),
            str(len(lr["vendor_urls"])),
            str(lr["markdown_len"]),
            status,
        )

    console.print(table)

    total_vendor_urls = sum(len(lr["vendor_urls"]) for lr in listing_results)
    console.print(f"  [bold]Total vendor profile URL kandidat:[/bold] [green]{total_vendor_urls}[/green]\n")

    # Print discovered URLs
    for lr in listing_results:
        if lr["vendor_urls"]:
            console.print(f"  [dim]From {lr['listing_url'][:60]}:[/dim]")
            for vu in lr["vendor_urls"][:5]:
                console.print(f"    [cyan]{vu}[/cyan]")
            if len(lr["vendor_urls"]) > 5:
                console.print(f"    [dim]... dan {len(lr['vendor_urls']) - 5} lainnya[/dim]")
            console.print()

    return listing_results

# ──────────────────────────────────────────────────────────────────────────────
# STEP 3 — Firecrawl: Scrape Vendor Profiles
# ──────────────────────────────────────────────────────────────────────────────

def step3_scrape_vendors(listing_results: list[dict]) -> list[dict]:
    """Scrape tiap vendor profile URL, extract structured data."""
    all_vendor_urls: list[tuple[str, str]] = []  # (vendor_url, listing_url)
    for lr in listing_results:
        for vu in lr["vendor_urls"]:
            all_vendor_urls.append((vu, lr["listing_url"]))

    console.print(Panel.fit(
        f"[bold green]STEP 3 — Firecrawl: Scrape Vendor Profiles[/bold green]\n"
        f"[dim]{len(all_vendor_urls)} profile URLs  "
        f"(max {MAX_PROFILES_PER_LISTING} per listing)[/dim]",
        border_style="green",
    ))

    if not all_vendor_urls:
        console.print("[red]  Tidak ada vendor profile URL ditemukan — cek Step 2[/red]")
        return []

    vendor_records: list[dict] = []
    seen_urls:      set[str]   = set()

    with Progress(
        SpinnerColumn(),
        TextColumn("[progress.description]{task.description:<62}"),
        TimeElapsedColumn(),
        console=console,
    ) as p:
        task = p.add_task("Scraping profiles ...", total=len(all_vendor_urls))

        for vendor_url, listing_url in all_vendor_urls:
            if vendor_url in seen_urls:
                p.advance(task)
                continue
            seen_urls.add(vendor_url)

            p.update(task, description=f"[green]{vendor_url[:60]}[/green]")
            fc = firecrawl_scrape(vendor_url, want_links=False, want_extract=True)
            time.sleep(1.0)

            extract = fc.get("extract", {}) or {}
            name    = (
                extract.get("company_name", "")
                or fc.get("title", "")
            ).strip()

            # Hitung confidence: berapa field terisi
            filled_fields = [
                k for k in ["company_name", "website", "email", "phone", "country",
                             "description", "booth_number", "category"]
                if extract.get(k)
            ]
            confidence = round(len(filled_fields) / 8.0, 2)

            vendor_records.append({
                "source_listing": listing_url,
                "source_url":     vendor_url,
                "name":           name,
                "website":        extract.get("website", ""),
                "email":          extract.get("email", ""),
                "phone":          extract.get("phone", ""),
                "country":        extract.get("country", ""),
                "city":           extract.get("city", ""),
                "description":    extract.get("description", "")[:300],
                "booth_number":   extract.get("booth_number", ""),
                "category":       extract.get("category", ""),
                "linkedin":       extract.get("linkedin", ""),
                "confidence":     confidence,
                "fields_filled":  len(filled_fields),
                "fc_success":     fc["success"],
                "fc_elapsed":     fc["elapsed"],
                "fc_error":       fc.get("error", ""),
                "markdown_len":   len(fc.get("markdown", "")),
                "markdown_500":   fc.get("markdown", "")[:500],
            })

            p.advance(task)

    # Summary table
    ok_vendors = [v for v in vendor_records if v["fc_success"] and v["name"]]

    table = Table(box=box.SIMPLE_HEAVY, header_style="bold white on dark_green", show_lines=True)
    table.add_column("Name",          min_width=35, style="green")
    table.add_column("Country",       width=12)
    table.add_column("Email",         width=28)
    table.add_column("Website",       width=30)
    table.add_column("Booth",         width=12)
    table.add_column("Confidence",    justify="center", width=12)
    table.add_column("Fields",        justify="center", width=8)

    for v in vendor_records:
        conf_color = "green" if v["confidence"] >= 0.5 else ("yellow" if v["confidence"] > 0 else "red")
        table.add_row(
            v["name"][:35] or "[dim](no name)[/dim]",
            v["country"][:12],
            v["email"][:28],
            v["website"][:30],
            v["booth_number"][:12],
            f"[{conf_color}]{v['confidence']:.2f}[/{conf_color}]",
            str(v["fields_filled"]),
        )

    console.print(table)
    console.print(
        f"\n  [bold]Total vendor profiles scraped:[/bold] [white]{len(vendor_records)}[/white]  "
        f"[bold]Dengan nama:[/bold] [green]{len(ok_vendors)}[/green]  "
        f"[bold]Confidence ≥ 0.5:[/bold] [green]{sum(1 for v in vendor_records if v['confidence'] >= 0.5)}[/green]\n"
    )

    return vendor_records

# ──────────────────────────────────────────────────────────────────────────────
# EXCEL EXPORT
# ──────────────────────────────────────────────────────────────────────────────

def write_summary_sheet(wb, query, tavily_candidates, listing_results, vendor_records) -> None:
    ws = wb.create_sheet(title="Summary", index=0)
    ws.merge_cells("A1:D1")
    c = ws["A1"]
    c.value     = f"Pipeline: Tavily → Firecrawl → Vendors  |  {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    c.font      = Font(bold=True, size=13, color="FFFFFF")
    c.fill      = NAVY
    c.alignment = CENTER
    ws.row_dimensions[1].height = 28

    rows = [
        ("Search Query",               query),
        ("", ""),
        ("STEP 1 — Tavily",            ""),
        ("  Candidates found",          str(len(tavily_candidates))),
        ("", ""),
        ("STEP 2 — Firecrawl Listings", ""),
        ("  Listing pages scraped",     str(len(listing_results))),
        ("  Total links discovered",    str(sum(lr["links_found"] for lr in listing_results))),
        ("  Vendor profile URLs found", str(sum(len(lr["vendor_urls"]) for lr in listing_results))),
        ("", ""),
        ("STEP 3 — Firecrawl Vendors",  ""),
        ("  Profiles scraped",          str(len(vendor_records))),
        ("  Vendors with name",         str(sum(1 for v in vendor_records if v["name"]))),
        ("  Confidence ≥ 0.5",         str(sum(1 for v in vendor_records if v["confidence"] >= 0.5))),
        ("  Avg fields filled",         str(round(
            sum(v["fields_filled"] for v in vendor_records) / max(1, len(vendor_records)), 1
        ))),
        ("", ""),
        ("Firecrawl credits used (est.)", str(len(listing_results) + len(vendor_records))),
    ]

    for row_idx, (k, v) in enumerate(rows, start=3):
        ws.cell(row=row_idx, column=1, value=k).font = Font(bold=bool(k and not k.startswith("  ")), size=11)
        ws.cell(row=row_idx, column=2, value=v).font  = Font(size=11)

    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 30


def write_vendor_sheet(wb, vendor_records: list[dict]) -> None:
    ws = wb.create_sheet(title="Vendors Extracted")
    cols = [
        ("No",              5),
        ("Name",           40),
        ("Website",        35),
        ("Email",          28),
        ("Phone",          18),
        ("Country",        15),
        ("City",           15),
        ("Booth",          14),
        ("Category",       25),
        ("LinkedIn",       35),
        ("Description",    55),
        ("Confidence",     12),
        ("Fields Filled",  13),
        ("Source URL",     60),
        ("Listing Page",   60),
        ("FC Error",       30),
    ]
    _write_header(ws, cols, fill=GREEN2)

    for i, v in enumerate(vendor_records, start=2):
        stripe = STRIPE2 if i % 2 == 0 else None
        values = [
            i - 1,
            v["name"],
            v["website"],
            v["email"],
            v["phone"],
            v["country"],
            v["city"],
            v["booth_number"],
            v["category"],
            v["linkedin"],
            v["description"],
            v["confidence"],
            v["fields_filled"],
            v["source_url"],
            v["source_listing"],
            v["fc_error"],
        ]
        for col, val in enumerate(values, start=1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.font      = BODY_FONT
            cell.alignment = TOP_WRAP
            if stripe:
                cell.fill = stripe
        ws.row_dimensions[i].height = 45


def write_tavily_sheet(wb, candidates: list[dict]) -> None:
    ws = wb.create_sheet(title="Step1 Tavily URLs")
    cols = [
        ("No",             5),
        ("Tavily Score",  13),
        ("Listing Score", 14),
        ("URL",           65),
        ("Title",         45),
        ("Snippet",       60),
    ]
    _write_header(ws, cols, fill=BLUE2)
    for i, c in enumerate(candidates, start=2):
        stripe = STRIPE if i % 2 == 0 else None
        values = [i-1, round(c["score"], 4), c["listing_score"], c["url"], c["title"], c["snippet"]]
        for col, val in enumerate(values, start=1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.font      = BODY_FONT
            cell.alignment = TOP_WRAP
            if stripe:
                cell.fill = stripe
        ws.row_dimensions[i].height = 35


def write_listing_sheet(wb, listing_results: list[dict]) -> None:
    ws = wb.create_sheet(title="Step2 Listings")
    cols = [
        ("No",              5),
        ("Listing URL",    65),
        ("Title",          45),
        ("Links Found",    13),
        ("Vendor URLs",    13),
        ("MD Len",          9),
        ("Elapsed (s)",    12),
        ("Status",         10),
        ("Error",          35),
        ("Vendor URLs List", 100),
    ]
    _write_header(ws, cols, fill=ORANGE)
    for i, lr in enumerate(listing_results, start=2):
        stripe = PatternFill("solid", fgColor="FCE4D6") if i % 2 == 0 else None
        values = [
            i-1,
            lr["listing_url"],
            lr["fc_title"] or lr["listing_title"],
            lr["links_found"],
            len(lr["vendor_urls"]),
            lr["markdown_len"],
            lr["fc_elapsed"],
            "✅" if lr["fc_success"] else "❌",
            lr["fc_error"],
            "\n".join(lr["vendor_urls"]),
        ]
        for col, val in enumerate(values, start=1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.font      = BODY_FONT
            cell.alignment = TOP_WRAP
            if stripe:
                cell.fill = stripe
        ws.row_dimensions[i].height = max(45, len(lr["vendor_urls"]) * 15)


def export_excel(query, tavily_candidates, listing_results, vendor_records) -> str:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    ts       = datetime.now().strftime("%Y%m%d_%H%M%S")
    filepath = OUTPUT_DIR / f"test_pipeline_{ts}.xlsx"

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    write_summary_sheet(wb, query, tavily_candidates, listing_results, vendor_records)
    write_vendor_sheet(wb, vendor_records)
    write_tavily_sheet(wb, tavily_candidates)
    write_listing_sheet(wb, listing_results)

    wb.save(str(filepath))
    return str(filepath.resolve())

# ──────────────────────────────────────────────────────────────────────────────
# MAIN
# ──────────────────────────────────────────────────────────────────────────────

def main() -> int:
    console.print(Panel.fit(
        "[bold white]Pipeline Test: Tavily → Firecrawl Listing → Firecrawl Vendors → Excel[/bold white]\n"
        "[dim]Zero OpenSERP  ·  Zero raw HTML parsing  ·  Firecrawl handles JS rendering[/dim]",
        border_style="white",
        padding=(0, 2),
    ))

    console.print(f"  TAVILY_API_KEY    : {'[green]✅ Set[/green]' if TAVILY_API_KEY    else '[red]❌ Missing[/red]'}")
    console.print(f"  FIRECRAWL_API_KEY : {'[green]✅ Set[/green]' if FIRECRAWL_API_KEY else '[red]❌ Missing[/red]'}")
    console.print(f"  Query             : [cyan]{SEARCH_QUERY}[/cyan]")
    console.print(f"  Max listings      : [yellow]{MAX_LISTING_PAGES}[/yellow]")
    console.print(f"  Max profiles/listing: [yellow]{MAX_PROFILES_PER_LISTING}[/yellow]")
    console.print(f"  Est. credits      : [yellow]~{MAX_LISTING_PAGES + MAX_LISTING_PAGES * MAX_PROFILES_PER_LISTING}[/yellow]\n")

    if not TAVILY_API_KEY or not FIRECRAWL_API_KEY:
        console.print("[bold red]Set TAVILY_API_KEY dan FIRECRAWL_API_KEY di .env[/bold red]")
        return 1

    # Run pipeline
    tavily_candidates = step1_tavily_search(SEARCH_QUERY)
    listing_results   = step2_scrape_listings(tavily_candidates)
    vendor_records    = step3_scrape_vendors(listing_results)

    # Export
    console.print("[dim]Exporting to Excel ...[/dim]")
    excel_path = export_excel(SEARCH_QUERY, tavily_candidates, listing_results, vendor_records)

    good = sum(1 for v in vendor_records if v["confidence"] >= 0.5)
    console.print(Panel.fit(
        f"[bold]Pipeline Selesai![/bold]\n\n"
        f"  Tavily URLs           : [cyan]{len(tavily_candidates)}[/cyan]\n"
        f"  Listing pages scraped : [yellow]{len(listing_results)}[/yellow]\n"
        f"  Vendor profiles       : [white]{len(vendor_records)}[/white]\n"
        f"  Confidence ≥ 0.5      : [green]{good}[/green]\n"
        f"  Firecrawl credits est.: [yellow]~{len(listing_results) + len(vendor_records)}[/yellow]\n\n"
        f"  Excel: [cyan]{excel_path}[/cyan]",
        border_style="green",
        padding=(0, 2),
    ))
    return 0


if __name__ == "__main__":
    sys.exit(main())
