import sys
import time
from datetime import datetime
from pathlib import Path
from urllib.parse import urlparse
from typing import Optional

import httpx
from loguru import logger
from rich.console import Console
from rich.progress import (
    Progress, SpinnerColumn, TextColumn,
    BarColumn, TaskProgressColumn, TimeElapsedColumn,
)
from rich.table import Table
from rich.panel import Panel
from rich import box

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

# ──────────────────────────────────────────────────────────────────────────────
# KONFIGURASI
# ──────────────────────────────────────────────────────────────────────────────

OPENSERP_BASE      = "http://localhost:7000"
OUTPUT_DIR         = Path("output")
RESULTS_PER_QUERY  = 20    # jumlah hasil per query ke OpenSERP
REQUEST_TIMEOUT    = 90    # detik — Yandex & Baidu bisa butuh 20-30s, kasih ruang
DELAY_BETWEEN_QUERIES = 1.2  # detik jeda antar query (hindari rate limit)

# Timeout per engine (override REQUEST_TIMEOUT untuk engine yang lambat)
ENGINE_TIMEOUT: dict[str, int] = {
    "google":     45,
    "baidu":      90,   # Baidu lambat karena Chromium + server jauh
    "bing":       45,
    "duckduckgo": 45,
}

console = Console()
REGIONS = [
    {
        "name": "Global (English)",
        "engine": "google",
        "flag": "🌐",
        "queries": [
            "defense technology exhibition exhibitors list 2026",
            "cybersecurity expo vendors directory 2026",
            "industrial trade show exhibitor companies 2026",
        ],
        "extra_params": {},
    },
    {
        "name": "China (Baidu)",
        "engine": "baidu",
        "flag": "🇨🇳",
        "queries": [
            "防务展览 参展商名录 2026",
            "网络安全博览会 参展企业 2026",
            "工业展览会 展商列表 2026",
        ],
        "extra_params": {},
    },
    {
        "name": "Russia (Google RU)",
        "engine": "google",
        "flag": "🇷🇺",
        "queries": [
            "оборонная выставка участники экспоненты 2026",
            "промышленная выставка поставщики компании 2026",
            "кибербезопасность форум участники 2026",
        ],
        "extra_params": {"gl": "ru", "hl": "ru"},
    },
    {
        "name": "Japan (Google JP)",
        "engine": "google",
        "flag": "🇯🇵",
        "queries": [
            "防衛展示会 出展企業 リスト 2026",
            "セキュリティ展 出展者 一覧 2026",
            "産業技術展示会 出展社 2026",
        ],
        "extra_params": {"gl": "jp", "hl": "ja"},
    },
    {
        "name": "Korea (Google KR)",
        "engine": "google",
        "flag": "🇰🇷",
        "queries": [
            "방위산업 전시회 참가업체 목록 2026",
            "사이버보안 박람회 참가기업 2026",
            "산업기술 전시회 출품업체 2026",
        ],
        "extra_params": {"gl": "kr", "hl": "ko"},
    },
    {
        "name": "India (Google IN)",
        "engine": "google",
        "flag": "🇮🇳",
        "queries": [
            "defense exhibition India exhibitors list 2026",
            "Aero India DefExpo exhibitor directory 2026",
            "industrial expo India vendors companies 2026",
        ],
        "extra_params": {"gl": "in", "hl": "en"},
    },
    {
        "name": "Middle East / Arabic (Google AE)",
        "engine": "google",
        "flag": "🇦🇪",
        "queries": [
            "IDEX DSEI MSPO defense exhibition exhibitors 2026",
            "معرض دفاعي شركات مشاركة قائمة 2026",
            "Middle East defense expo vendor companies 2026",
        ],
        "extra_params": {"gl": "ae", "hl": "ar"},
    },
    {
        "name": "Pakistan (Google PK)",
        "engine": "google",
        "flag": "🇵🇰",
        "queries": [
            "IDEAS Karachi defense exhibition exhibitors 2026",
            "Pakistan industrial expo vendors list 2026",
        ],
        "extra_params": {"gl": "pk", "hl": "en"},
    },
]

# ──────────────────────────────────────────────────────────────────────────────
# LOGGING
# ──────────────────────────────────────────────────────────────────────────────

def setup_logging() -> Path:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    log_path = OUTPUT_DIR / "test_openserp.log"

    logger.remove()

    # Terminal: INFO saja — WARNING/ERROR tidak muncul di console, hanya ke file
    logger.add(
        sys.stderr,
        format="<green>{time:HH:mm:ss}</green> | <cyan>{message}</cyan>",
        level="INFO",
        filter=lambda record: record["level"].name == "INFO",
        colorize=True,
    )

    # File: DEBUG ke atas, format lengkap — semua error/warning tersimpan di sini
    logger.add(
        str(log_path),
        format=(
            "{time:YYYY-MM-DD HH:mm:ss.SSS} | "
            "{level: <8} | "
            "{name}:{function}:{line} | "
            "{message}"
        ),
        level="DEBUG",
        rotation="10 MB",
        encoding="utf-8",
        enqueue=True,
    )

    return log_path

# ──────────────────────────────────────────────────────────────────────────────
# OPENSERP CLIENT
# ──────────────────────────────────────────────────────────────────────────────

def check_openserp_health() -> bool:
    """Cek apakah OpenSERP container sudah berjalan."""
    logger.info(f"Checking OpenSERP at {OPENSERP_BASE} ...")
    try:
        r = httpx.get(f"{OPENSERP_BASE}/health", timeout=5)
        if r.status_code == 200:
            logger.info("OpenSERP health check: OK ✓")
            return True
        logger.error(f"OpenSERP health returned HTTP {r.status_code}")
        return False
    except httpx.ConnectError:
        logger.error(
            "Tidak bisa connect ke OpenSERP. Pastikan Docker container sudah jalan:\n"
            "  docker run -p 127.0.0.1:7000:7000 -it karust/openserp serve -a 0.0.0.0 -p 7000"
        )
        return False
    except Exception as e:
        logger.error(f"Health check exception: {e}")
        return False


# Endpoint pattern yang berhasil, di-discover sekali saat startup
_ENDPOINT_PATTERN: Optional[str] = None  # "new" | "old"


def discover_endpoint_pattern() -> str:
    """
    OpenSERP versi berbeda pakai format endpoint yang berbeda.
    Coba semua kemungkinan dengan query dummy, simpan yang berhasil.

    Format yang dicoba (urutan prioritas):
      1. GET /search?engine=google&q=test          (v0.6+)
      2. GET /google?q=test                        (v0.5.x, engine sebagai path)
      3. GET /google?text=test                     (varian lama, param = text)
    """
    global _ENDPOINT_PATTERN

    if _ENDPOINT_PATTERN is not None:
        return _ENDPOINT_PATTERN

    candidates = [
        # Format confirmed: /{engine}/search?text=...
        ("engine_text",  f"{OPENSERP_BASE}/google/search",  {"text": "test", "limit": 1}),
        # Fallback dengan q= kalau versi lain
        ("engine_q",     f"{OPENSERP_BASE}/google/search",  {"q": "test", "limit": 1}),
    ]

    for pattern, url, params in candidates:
        try:
            r = httpx.get(url, params=params, timeout=10)
            logger.debug(f"  endpoint probe [{pattern}] → HTTP {r.status_code}  body={r.text[:200]!r}")
            if r.status_code == 200:
                logger.info(f"Endpoint pattern detected: [{pattern}]  ({url})")
                _ENDPOINT_PATTERN = pattern
                return pattern
        except Exception as e:
            logger.debug(f"  endpoint probe [{pattern}] failed: {e}")

    logger.warning("Tidak bisa detect endpoint pattern — fallback ke 'old_q'")
    _ENDPOINT_PATTERN = "old_q"
    return "old_q"


def _build_request(pattern: str, engine: str, query: str, limit: int, extra: dict) -> tuple[str, dict]:
    """Bangun URL + params sesuai endpoint pattern yang aktif."""
    url = f"{OPENSERP_BASE}/{engine}/search"
    if pattern == "engine_q":
        params = {"q": query, "limit": limit}
    else:  # engine_text (default/confirmed)
        params = {"text": query, "limit": limit}

    params.update(extra)
    return url, params


def _parse_response(data) -> list[dict]:
    """Normalisasi berbagai bentuk respons JSON dari OpenSERP."""
    if isinstance(data, list):
        return data
    if isinstance(data, dict):
        for key in ("results", "data", "items", "organic"):
            if key in data and isinstance(data[key], list):
                return data[key]
    return []


def search_one(
    engine: str,
    query: str,
    limit: int = RESULTS_PER_QUERY,
    extra_params: Optional[dict] = None,
) -> list[dict]:
    """Satu panggilan search ke OpenSERP. Return list hasil mentah."""
    pattern = discover_endpoint_pattern()
    url, params = _build_request(pattern, engine, query, limit, extra_params or {})
    timeout = ENGINE_TIMEOUT.get(engine, REQUEST_TIMEOUT)

    logger.debug(f"→ SEARCH [{engine.upper()}] {query!r} | pattern={pattern} timeout={timeout}s extra={extra_params}")

    try:
        t0      = time.perf_counter()
        resp    = httpx.get(url, params=params, timeout=timeout)
        elapsed = time.perf_counter() - t0

        if resp.status_code != 200:
            logger.warning(f"  [{engine}] HTTP {resp.status_code} — query={query!r}")
            logger.debug(f"  [{engine}] body: {resp.text[:300]}")
            return []

        results = _parse_response(resp.json())
        logger.info(f"  [{engine}] {len(results):>3} hasil  ({elapsed:.1f}s)  — {query[:55]!r}")
        return results

    except httpx.TimeoutException:
        logger.warning(f"  [{engine}] Timeout setelah {timeout}s — {query!r}")
        return []
    except Exception as e:
        logger.error(f"  [{engine}] Error: {e} — {query!r}")
        return []

# ──────────────────────────────────────────────────────────────────────────────
# NORMALIZER
# ──────────────────────────────────────────────────────────────────────────────

def extract_domain(url: str) -> str:
    try:
        return urlparse(url).netloc.lstrip("www.")
    except Exception:
        return ""


def normalize_result(raw: dict, region: str, engine: str, query: str) -> Optional[dict]:
    """
    Ubah satu hasil mentah OpenSERP jadi format yang konsisten.
    OpenSERP mungkin pakai field name berbeda-beda tergantung engine.
    """
    # Coba berbagai kemungkinan field name
    url     = raw.get("url") or raw.get("link") or raw.get("href") or ""
    title   = raw.get("title") or raw.get("name") or raw.get("header") or ""
    snippet = (
        raw.get("description")
        or raw.get("snippet")
        or raw.get("text")
        or raw.get("body")
        or ""
    )

    if not url or not title:
        return None

    return {
        "region":     region,
        "engine":     engine,
        "title":      title.strip(),
        "url":        url.strip(),
        "domain":     extract_domain(url),
        "snippet":    snippet.strip()[:350],
        "query_used": query,
        "fetched_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    }


def deduplicate(records: list[dict]) -> list[dict]:
    """Hapus URL duplikat dalam satu region."""
    seen: set[str] = set()
    out: list[dict] = []
    for r in records:
        if r["url"] not in seen:
            seen.add(r["url"])
            out.append(r)
    return out

# ──────────────────────────────────────────────────────────────────────────────
# EXCEL EXPORT
# ──────────────────────────────────────────────────────────────────────────────

NAVY   = PatternFill("solid", fgColor="1F3864")
BLUE2  = PatternFill("solid", fgColor="2E75B6")
STRIPE = PatternFill("solid", fgColor="DCE6F1")

HDR_FONT  = Font(bold=True, color="FFFFFF", size=11)
BODY_FONT = Font(size=10)
CENTER    = Alignment(horizontal="center", vertical="center", wrap_text=True)
TOP_WRAP  = Alignment(vertical="top", wrap_text=True)

SHEET_COLUMNS = [
    ("No",         5),
    ("Region",    22),
    ("Engine",    12),
    ("Title",     50),
    ("URL",       55),
    ("Domain",    30),
    ("Snippet",   65),
    ("Query Used",42),
    ("Fetched At",20),
]


def _write_header(ws) -> None:
    for col, (label, width) in enumerate(SHEET_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col, value=label)
        cell.fill   = NAVY
        cell.font   = HDR_FONT
        cell.alignment = CENTER
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"


def _write_region_sheet(wb: openpyxl.Workbook, sheet_name: str, records: list[dict]) -> None:
    ws = wb.create_sheet(title=sheet_name[:31])
    _write_header(ws)

    for row_idx, rec in enumerate(records, start=2):
        stripe = STRIPE if row_idx % 2 == 0 else None
        values = [
            row_idx - 1,
            rec["region"],
            rec["engine"],
            rec["title"],
            rec["url"],
            rec["domain"],
            rec["snippet"],
            rec["query_used"],
            rec["fetched_at"],
        ]
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font      = BODY_FONT
            cell.alignment = TOP_WRAP
            if stripe:
                cell.fill = stripe
        ws.row_dimensions[row_idx].height = 45


def _write_summary_sheet(wb: openpyxl.Workbook, stats: list[dict], log_path: str) -> None:
    ws = wb.create_sheet(title="Summary", index=0)

    # Judul besar
    ws.merge_cells("A1:F1")
    c = ws["A1"]
    c.value     = f"OpenSERP Multi-Region Test — {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    c.font      = Font(bold=True, size=14, color="FFFFFF")
    c.fill      = NAVY
    c.alignment = CENTER
    ws.row_dimensions[1].height = 30

    # Sub-judul: log path
    ws.merge_cells("A2:F2")
    c2 = ws["A2"]
    c2.value     = f"Log: {log_path}"
    c2.font      = Font(italic=True, size=10, color="595959")
    c2.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 16

    # Header tabel
    sum_headers = ["Region", "Engine", "Flag", "Queries Run", "Unique Results", "Status"]
    sum_widths  = [28, 14, 8, 15, 18, 14]
    for col, (h, w) in enumerate(zip(sum_headers, sum_widths), start=1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.fill      = BLUE2
        cell.font      = HDR_FONT
        cell.alignment = CENTER
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[3].height = 20

    # Baris data
    for row_idx, stat in enumerate(stats, start=4):
        n      = stat["results"]
        status = "✅ OK" if n >= 20 else ("⚠️ Low (<20)" if n > 0 else "❌ Failed")
        stripe = STRIPE if row_idx % 2 == 0 else None
        values = [stat["region"], stat["engine"], stat["flag"],
                  stat["queries_run"], n, status]
        for col, val in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.font      = BODY_FONT
            cell.alignment = CENTER
            if stripe:
                cell.fill = stripe

    ws.freeze_panes = "A4"


def export_excel(
    all_results: dict[str, list[dict]],
    stats: list[dict],
    log_path: str,
) -> str:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    ts       = datetime.now().strftime("%Y%m%d_%H%M%S")
    filepath = OUTPUT_DIR / f"test_openserp_{ts}.xlsx"

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # Sheet 1: Summary
    _write_summary_sheet(wb, stats, log_path)

    # Sheet per region — strip karakter yang tidak valid di nama sheet Excel
    _invalid = r'\/:?*[]'
    def _safe_sheet_name(name: str) -> str:
        for ch in _invalid:
            name = name.replace(ch, "-")
        return name.split("(")[0].strip()[:31]

    for region_name, records in all_results.items():
        _write_region_sheet(wb, _safe_sheet_name(region_name), records)

    # Sheet terakhir: semua hasil gabungan
    all_flat = [r for records in all_results.values() for r in records]
    if all_flat:
        _write_region_sheet(wb, "All Results", all_flat)

    wb.save(str(filepath))
    logger.info(f"Excel saved: {filepath.resolve()}")
    return str(filepath.resolve())

# ──────────────────────────────────────────────────────────────────────────────
# MAIN
# ──────────────────────────────────────────────────────────────────────────────

def main() -> int:
    log_path = setup_logging()

    console.print(Panel.fit(
        "[bold cyan]OpenSERP Multi-Region Vendor Search Test[/bold cyan]\n"
        f"[dim]Target: {OPENSERP_BASE}  |  Regions: {len(REGIONS)}  |  {RESULTS_PER_QUERY} results/query[/dim]",
        border_style="cyan",
        padding=(0, 2),
    ))

    logger.info("=" * 70)
    logger.info("OpenSERP Multi-Region Test START")
    logger.info(f"Base URL  : {OPENSERP_BASE}")
    logger.info(f"Regions   : {len(REGIONS)}")
    logger.info(f"Per query : {RESULTS_PER_QUERY} results")
    logger.info(f"Log file  : {log_path.resolve()}")
    logger.info("=" * 70)

    # ── Health check ──────────────────────────────────────────────────────────
    if not check_openserp_health():
        console.print(
            "[bold red]OpenSERP tidak berjalan.[/bold red] "
            "Jalankan dulu:\n"
            "  [yellow]docker run -p 127.0.0.1:7000:7000 -it karust/openserp "
            "serve -a 0.0.0.0 -p 7000[/yellow]"
        )
        return 1

    all_results: dict[str, list[dict]] = {}
    region_stats: list[dict] = []

    # ── Search loop ───────────────────────────────────────────────────────────
    with Progress(
        SpinnerColumn(),
        TextColumn("[progress.description]{task.description:<40}"),
        BarColumn(bar_width=30),
        TaskProgressColumn(),
        TimeElapsedColumn(),
        console=console,
        transient=False,
    ) as progress:

        overall = progress.add_task("[bold white]Total progress", total=len(REGIONS))

        for region in REGIONS:
            name     = region["name"]
            engine   = region["engine"]
            flag     = region["flag"]
            queries  = region["queries"]
            extra    = region.get("extra_params", {})

            region_task = progress.add_task(
                f"  {flag} [cyan]{name}[/cyan]",
                total=len(queries),
            )

            logger.info(f"── {flag}  Region: {name}  [{engine.upper()}] ──────")

            raw_records: list[dict] = []
            queries_run = 0

            for query in queries:
                raw_list = search_one(engine, query, limit=RESULTS_PER_QUERY, extra_params=extra)

                for raw in raw_list:
                    rec = normalize_result(raw, name, engine, query)
                    if rec:
                        raw_records.append(rec)

                queries_run += 1
                progress.advance(region_task)
                time.sleep(DELAY_BETWEEN_QUERIES)

            deduped = deduplicate(raw_records)
            all_results[name] = deduped

            status_log = "✅" if len(deduped) >= 20 else ("⚠️ " if len(deduped) > 0 else "❌")
            logger.info(f"  {status_log} {name}: {len(deduped)} unique results (dari {len(raw_records)} raw)")

            region_stats.append({
                "region":      name,
                "engine":      engine,
                "flag":        flag,
                "queries_run": queries_run,
                "results":     len(deduped),
            })

            progress.remove_task(region_task)
            progress.advance(overall)

    # ── Export Excel ──────────────────────────────────────────────────────────
    logger.info("Exporting to Excel ...")
    excel_path = export_excel(all_results, region_stats, str(log_path.resolve()))

    # ── Summary table di terminal ─────────────────────────────────────────────
    table = Table(
        title="[bold]Hasil Per Region[/bold]",
        box=box.ROUNDED,
        border_style="cyan",
        header_style="bold white on dark_blue",
        show_lines=True,
    )
    table.add_column("Flag",    justify="center", width=5)
    table.add_column("Region",  style="cyan",     min_width=28)
    table.add_column("Engine",  style="yellow",   justify="center", width=12)
    table.add_column("Queries", justify="center", width=8)
    table.add_column("Results", justify="center", width=9)
    table.add_column("Status",  justify="center", width=14)

    total_results = 0
    for stat in region_stats:
        n = stat["results"]
        total_results += n
        if n >= 20:
            status_str = "[green]✅ OK[/green]"
        elif n > 0:
            status_str = "[yellow]⚠️  Low[/yellow]"
        else:
            status_str = "[red]❌ Failed[/red]"
        table.add_row(
            stat["flag"],
            stat["region"],
            stat["engine"],
            str(stat["queries_run"]),
            str(n),
            status_str,
        )

    console.print()
    console.print(table)
    console.print()
    console.print(f"[bold]Total unique results :[/bold] [green]{total_results}[/green]")
    console.print(f"[bold]Excel output         :[/bold] [cyan]{excel_path}[/cyan]")
    console.print(f"[bold]Log file             :[/bold] [cyan]{log_path.resolve()}[/cyan]")
    console.print()

    logger.info("=" * 70)
    logger.info(f"Test COMPLETE — {total_results} total results across {len(REGIONS)} regions")
    logger.info(f"Excel  : {excel_path}")
    logger.info(f"Log    : {log_path.resolve()}")
    logger.info("=" * 70)

    failed_regions = [s["region"] for s in region_stats if s["results"] == 0]
    if failed_regions:
        logger.warning(f"Region dengan 0 hasil: {failed_regions}")
        console.print(
            f"[yellow]⚠️  {len(failed_regions)} region tidak menghasilkan data.[/yellow] "
            "Cek log untuk detail error."
        )

    return 0


if __name__ == "__main__":
    sys.exit(main())
