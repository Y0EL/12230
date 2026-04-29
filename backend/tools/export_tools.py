import re
from datetime import datetime
from pathlib import Path
from typing import Optional

import pandas as pd
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.series import DataPoint
from langchain_core.tools import tool
from loguru import logger

from backend.core.config import get_settings

_settings = get_settings()

VENDOR_COLUMNS = [
    ("name",           "Company Name",       35),
    ("website",        "Website",            30),
    ("email",          "Email",              28),
    ("phone",          "Phone",              18),
    ("address",        "Address",            35),
    ("city",           "City",               18),
    ("country",        "Country",            15),
    ("category",       "Category",           22),
    ("description",    "Description",        50),
    ("linkedin",       "LinkedIn",           30),
    ("twitter",        "Twitter / X",        25),
    ("booth_number",   "Booth / Stand",      14),
    ("event_name",     "Event Name",         30),
    ("event_location", "Event Location",     22),
    ("event_date",     "Event Date",         15),
    ("source_url",     "Source URL",         35),
    ("extraction_method", "Extract Method",  15),
    ("confidence_score",  "Confidence",       10),
]

COLOR_HEADER_BG   = "1E3A5F"
COLOR_HEADER_FONT = "FFFFFF"
COLOR_ROW_ODD     = "F0F5FB"
COLOR_ROW_EVEN    = "FFFFFF"
COLOR_SCHEMA_ORG  = "C6EFCE"
COLOR_RULE_BASED  = "FFEB9C"
COLOR_LLM         = "FFC7CE"
COLOR_ACCENT      = "2E75B6"
COLOR_BORDER      = "BDD7EE"

THIN_BORDER = Border(
    left=Side(style="thin", color=COLOR_BORDER),
    right=Side(style="thin", color=COLOR_BORDER),
    top=Side(style="thin", color=COLOR_BORDER),
    bottom=Side(style="thin", color=COLOR_BORDER),
)

MEDIUM_BORDER = Border(
    left=Side(style="medium", color=COLOR_ACCENT),
    right=Side(style="medium", color=COLOR_ACCENT),
    top=Side(style="medium", color=COLOR_ACCENT),
    bottom=Side(style="medium", color=COLOR_ACCENT),
)


def _make_header_style() -> tuple:
    font = Font(name="Calibri", bold=True, color=COLOR_HEADER_FONT, size=11)
    fill = PatternFill(fill_type="solid", fgColor=COLOR_HEADER_BG)
    align = Alignment(horizontal="center", vertical="center", wrap_text=False)
    return font, fill, align


def _make_cell_style(row_idx: int, extraction_method: str = "") -> tuple:
    if extraction_method == "schema_org":
        bg = COLOR_SCHEMA_ORG
    elif extraction_method == "rule_based":
        bg = COLOR_RULE_BASED
    elif "llm" in extraction_method:
        bg = COLOR_LLM
    elif row_idx % 2 == 0:
        bg = COLOR_ROW_EVEN
    else:
        bg = COLOR_ROW_ODD

    font = Font(name="Calibri", size=10)
    fill = PatternFill(fill_type="solid", fgColor=bg)
    align = Alignment(vertical="top", wrap_text=False)
    return font, fill, align


def _sanitize_cell(value) -> str:
    if value is None:
        return ""
    val = str(value).strip()
    val = re.sub(r'[\x00-\x08\x0b\x0c\x0e-\x1f\x7f]', '', val)
    return val[:32767]


def _build_vendors_sheet(ws, vendors: list[dict]) -> None:
    h_font, h_fill, h_align = _make_header_style()

    ws.row_dimensions[1].height = 22

    for col_idx, (field, header, width) in enumerate(VENDOR_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = h_font
        cell.fill = h_fill
        cell.alignment = h_align
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(VENDOR_COLUMNS))}1"

    for row_idx, vendor in enumerate(vendors, start=2):
        method = vendor.get("extraction_method", "")
        v_font, v_fill, v_align = _make_cell_style(row_idx, method)
        ws.row_dimensions[row_idx].height = 16

        for col_idx, (field, _, _) in enumerate(VENDOR_COLUMNS, start=1):
            raw_val = vendor.get(field, "")
            if field == "confidence_score" and raw_val:
                try:
                    raw_val = f"{float(raw_val) * 100:.0f}%"
                except (ValueError, TypeError):
                    raw_val = str(raw_val)
            cell = ws.cell(row=row_idx, column=col_idx, value=_sanitize_cell(raw_val))
            cell.font = v_font
            cell.fill = v_fill
            cell.alignment = v_align
            cell.border = THIN_BORDER

            if field in ("website", "linkedin", "twitter", "source_url"):
                url_val = vendor.get(field, "")
                if url_val and url_val.startswith("http"):
                    cell.hyperlink = url_val
                    cell.font = Font(
                        name="Calibri", size=10, color="0563C1", underline="single"
                    )


def _build_stats_sheet(ws, vendors: list[dict], query: str, stats: dict) -> None:
    h_font, h_fill, h_align = _make_header_style()

    ws["A1"] = "MEGA CRAWLER BOT — Run Summary"
    ws["A1"].font = Font(name="Calibri", bold=True, size=14, color=COLOR_ACCENT)
    ws.row_dimensions[1].height = 28

    ws["A3"] = "Query"
    ws["B3"] = query
    ws["A4"] = "Run Date"
    ws["B4"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws["A5"] = "Total Vendors"
    ws["B5"] = len(vendors)
    ws["A5"].font = Font(bold=True)
    ws["B5"].font = Font(bold=True, color="375623")

    label_font = Font(name="Calibri", bold=True, size=10)
    for row in [3, 4, 5]:
        ws.cell(row=row, column=1).font = label_font

    stat_rows = [
        ("Pages Crawled", stats.get("total_crawled", 0)),
        ("Vendor Pages Found", stats.get("total_vendor_pages", 0)),
        ("via schema.org", stats.get("extraction_schema_org", 0)),
        ("via rule_based", stats.get("extraction_rule_based", 0)),
        ("via LLM fallback", stats.get("extraction_llm", 0)),
        ("Extraction Failed", stats.get("extraction_failed", 0)),
        ("LLM Usage %", f"{stats.get('llm_percentage', 0):.1f}%"),
        ("Est. LLM Cost (USD)", f"${stats.get('estimated_llm_cost_usd', 0):.4f}"),
        ("Countries Found", stats.get("countries_found", 0)),
        ("Events Found", stats.get("events_found", 0)),
        ("Success Rate", f"{stats.get('success_rate', 0):.1f}%"),
        ("Elapsed (sec)", stats.get("elapsed_seconds", 0)),
    ]

    row_start = 7
    ws.cell(row=row_start, column=1, value="Metric").font = h_font
    ws.cell(row=row_start, column=1).fill = h_fill
    ws.cell(row=row_start, column=2, value="Value").font = h_font
    ws.cell(row=row_start, column=2).fill = h_fill

    for i, (metric, value) in enumerate(stat_rows, start=row_start + 1):
        ws.cell(row=i, column=1, value=metric)
        ws.cell(row=i, column=2, value=value)
        if i % 2 == 0:
            for c in [1, 2]:
                ws.cell(row=i, column=c).fill = PatternFill(fill_type="solid", fgColor=COLOR_ROW_ODD)

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 20

    country_start = row_start + len(stat_rows) + 3
    ws.cell(row=country_start, column=1, value="Top Countries").font = Font(bold=True, size=11)
    from collections import Counter
    country_counts = Counter(v.get("country", "Unknown") or "Unknown" for v in vendors)
    for i, (country, count) in enumerate(country_counts.most_common(15), start=country_start + 1):
        ws.cell(row=i, column=1, value=country or "Unknown")
        ws.cell(row=i, column=2, value=count)

    cat_start = country_start
    ws.cell(row=cat_start, column=4, value="Top Categories").font = Font(bold=True, size=11)
    cat_counts = Counter(v.get("category", "Unknown") or "Unknown" for v in vendors)
    for i, (cat, count) in enumerate(cat_counts.most_common(15), start=cat_start + 1):
        ws.cell(row=i, column=4, value=cat or "Unknown")
        ws.cell(row=i, column=5, value=count)

    ws.column_dimensions["D"].width = 25
    ws.column_dimensions["E"].width = 10


def _build_dedup_check_sheet(ws, vendors: list[dict]) -> None:
    h_font, h_fill, _ = _make_header_style()

    headers = ["Company Name", "Website Domain", "Country", "Count"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = h_font
        cell.fill = h_fill
        ws.column_dimensions[get_column_letter(col)].width = 28

    from collections import Counter
    from urllib.parse import urlparse

    def get_domain(url: str) -> str:
        if not url:
            return ""
        try:
            return urlparse(url).netloc.lower()
        except Exception:
            return ""

    name_counts = Counter(
        (v.get("name", "") or "").lower().strip() for v in vendors if v.get("name")
    )
    domain_counts = Counter(
        get_domain(v.get("website", "")) for v in vendors if v.get("website")
    )

    duplicates = [(k, v) for k, v in name_counts.items() if v > 1]
    duplicates.sort(key=lambda x: x[1], reverse=True)

    for row_idx, (name, count) in enumerate(duplicates[:100], start=2):
        matching = [v for v in vendors if (v.get("name", "") or "").lower().strip() == name]
        first = matching[0] if matching else {}
        ws.cell(row=row_idx, column=1, value=name[:100])
        ws.cell(row=row_idx, column=2, value=get_domain(first.get("website", "")))
        ws.cell(row=row_idx, column=3, value=first.get("country", ""))
        ws.cell(row=row_idx, column=4, value=count)
        if count > 1:
            for col in range(1, 5):
                ws.cell(row=row_idx, column=col).fill = PatternFill(fill_type="solid", fgColor="FFC7CE")


@tool
def export_to_excel(vendors: list[dict] = [], query: str = "", stats: dict = None) -> str:
    """
    Export vendor list to a formatted Excel file with multiple sheets:
    Sheet 1: All vendors with colored rows by extraction method
    Sheet 2: Run statistics and country/category breakdown
    Sheet 3: Duplicate check

    If vendors list is empty (or not provided), exports from the global registry.
    Always pass query = the original user query string.

    Returns the absolute file path of the created Excel file.
    """
    from backend.tools.vendor_registry import get_all_vendors
    settings = get_settings()

    # Prefer registry over passed vendors (registry always has the full, deduped set)
    registry = get_all_vendors()
    effective_vendors = registry if registry else vendors

    if not effective_vendors:
        logger.warning("No vendors to export (registry empty and no vendors passed)")
        return ""

    vendors = effective_vendors

    stats = stats or {}
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    safe_query = re.sub(r'[^\w\s-]', '', query)[:30].strip().replace(" ", "_")
    filename = f"vendors_{safe_query}_{timestamp}.xlsx"
    output_path = settings.output_path / filename

    wb = openpyxl.Workbook()
    wb.properties.title = f"Vendor List — {query}"
    wb.properties.creator = "Mega Crawler Bot"
    wb.properties.description = f"Generated {datetime.now().isoformat()}"

    ws_vendors = wb.active
    ws_vendors.title = "Vendors"
    ws_vendors.sheet_view.showGridLines = True
    ws_vendors.sheet_properties.tabColor = "1E3A5F"
    _build_vendors_sheet(ws_vendors, vendors)

    ws_stats = wb.create_sheet("Statistics")
    ws_stats.sheet_properties.tabColor = "375623"
    _build_stats_sheet(ws_stats, vendors, query, stats)

    ws_dedup = wb.create_sheet("Duplicates Check")
    ws_dedup.sheet_properties.tabColor = "C00000"
    _build_dedup_check_sheet(ws_dedup, vendors)

    wb.active = wb["Vendors"]

    wb.save(str(output_path))
    logger.info(f"Excel exported: {output_path} ({len(vendors)} vendors)")
    return str(output_path)


@tool
def export_to_csv(vendors: list[dict] = [], query: str = "") -> str:
    """
    Export vendor list to a CSV file.
    If vendors list is empty (or not provided), exports from the global registry.
    Always pass query = the original user query string.
    Returns the absolute file path of the created CSV file.
    """
    from backend.tools.vendor_registry import get_all_vendors
    settings = get_settings()

    # Prefer registry over passed vendors
    registry = get_all_vendors()
    effective_vendors = registry if registry else vendors

    if not effective_vendors:
        return ""

    vendors = effective_vendors

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    safe_query = re.sub(r'[^\w\s-]', '', query)[:30].strip().replace(" ", "_")
    filename = f"vendors_{safe_query}_{timestamp}.csv"
    output_path = settings.output_path / filename

    columns = [field for field, _, _ in VENDOR_COLUMNS]
    rows = []
    for vendor in vendors:
        row = {col: _sanitize_cell(vendor.get(col, "")) for col in columns}
        rows.append(row)

    df = pd.DataFrame(rows, columns=columns)
    df.to_csv(str(output_path), index=False, encoding="utf-8-sig")
    logger.info(f"CSV exported: {output_path} ({len(vendors)} vendors)")
    return str(output_path)


def _do_deduplicate(vendors: list[dict]) -> list[dict]:
    """
    Internal deduplication logic.
    Remove duplicates based on name similarity and website domain.
    Keeps the record with highest confidence_score and merges fields.
    """
    from urllib.parse import urlparse
    from collections import defaultdict

    if not vendors:
        return []

    def normalize_name(name: str) -> str:
        if not name:
            return ""
        name = name.lower().strip()
        name = re.sub(r'\b(inc|llc|ltd|corp|corporation|company|co|gmbh|ag|bv|sa|spa|plc|pty|nv)\b', '', name)
        name = re.sub(r'[^a-z0-9\s]', '', name)
        name = re.sub(r'\s+', ' ', name).strip()
        return name

    def get_domain(url: str) -> str:
        if not url:
            return ""
        try:
            parsed = urlparse(url)
            domain = parsed.netloc.lower()
            return domain.replace("www.", "")
        except Exception:
            return ""

    groups: dict[str, list[dict]] = defaultdict(list)

    for vendor in vendors:
        name_key = normalize_name(vendor.get("name", ""))
        domain_key = get_domain(vendor.get("website", ""))
        key = domain_key if domain_key else name_key
        if not key:
            key = f"__unnamed_{id(vendor)}"
        groups[key].append(vendor)

    deduped = []
    for key, group in groups.items():
        if len(group) == 1:
            deduped.append(group[0])
        else:
            best = max(group, key=lambda v: v.get("confidence_score", 0))
            for other in group:
                if other is not best:
                    for field, _, _ in VENDOR_COLUMNS:
                        if not best.get(field) and other.get(field):
                            best[field] = other[field]
            deduped.append(best)

    deduped.sort(key=lambda v: v.get("confidence_score", 0), reverse=True)
    return deduped


@tool
def deduplicate_vendors(vendors: list[dict] = []) -> dict:
    """
    Deduplicate all collected vendors.

    If vendors list is empty (or not provided), operates on the global registry
    (which contains ALL vendors extracted during this run). This is the recommended
    usage — call without arguments after crawling is done.

    After deduplication, the registry is updated with the clean list so that
    export_to_excel and export_to_csv will use the deduplicated data.

    Returns a summary dict (NOT the full list, to avoid context overflow):
      {"original_count": N, "deduped_count": M, "message": "..."}
    """
    from backend.tools.vendor_registry import get_all_vendors, replace_all

    # Prefer registry (full set) over whatever partial list agent might pass
    registry = get_all_vendors()
    source = registry if registry else vendors

    if not source:
        return {"original_count": 0, "deduped_count": 0, "message": "No vendors to deduplicate."}

    deduped = _do_deduplicate(source)

    # Update the registry with the deduplicated list so exports use clean data
    replace_all(deduped)

    msg = (
        f"Deduplication complete: {len(source)} -> {len(deduped)} unique vendors. "
        f"Registry updated. Call export_to_excel(query='...') to save."
    )
    logger.info(f"Deduplicated: {len(source)} -> {len(deduped)} vendors (registry updated)")
    return {"original_count": len(source), "deduped_count": len(deduped), "message": msg}


ALL_EXPORT_TOOLS = [export_to_excel, export_to_csv, deduplicate_vendors]
