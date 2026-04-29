"""
Test Tavily Search + Firecrawl Scrape — standalone, no LLM, results saved to Excel.

Usage:
    python test_firecrawl_tavily.py

Keys dibaca dari .env di root project:
    TAVILY_API_KEY=tvly-xxxxx
    FIRECRAWL_API_KEY=fc-xxxxx
"""

import sys
import io
import os
import time
import json
from datetime import datetime
from pathlib import Path
from urllib.parse import urlparse

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

import httpx
from rich.console import Console
from rich.table import Table
from rich.panel import Panel
from rich.progress import (
    Progress, SpinnerColumn, TextColumn,
    BarColumn, TaskProgressColumn, TimeElapsedColumn,
)
from rich import box
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

# ──────────────────────────────────────────────────────────────────────────────
# Load .env
# ──────────────────────────────────────────────────────────────────────────────

def _load_env() -> None:
    env_path = Path(__file__).resolve().parent / ".env"
    if not env_path.exists():
        return
    with open(env_path, encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            key, _, val = line.partition("=")
            key = key.strip()
            val = val.strip().strip('"').strip("'")
            if key and key not in os.environ:
                os.environ[key] = val

_load_env()

TAVILY_API_KEY    = os.environ.get("TAVILY_API_KEY", "")
FIRECRAWL_API_KEY = os.environ.get("FIRECRAWL_API_KEY", "")

OUTPUT_DIR = Path("output")
console    = Console()

# ──────────────────────────────────────────────────────────────────────────────
# Queries untuk Tavily
# ──────────────────────────────────────────────────────────────────────────────

TAVILY_QUERIES = [
    "defense technology exhibition exhibitors list 2026",
    "cybersecurity expo vendors directory 2025 2026",
    "military aerospace trade show exhibitor companies",
    "DSEI IDEX DSEi exhibitors list",
    "industrial technology fair participants directory",
    "security conference sponsors vendors list",
]

# URLs yang langsung di-scrape Firecrawl (mix: vendor profile + listing page)
FIRECRAWL_TEST_URLS = [
    "https://www.sahaexpo.com/en/exhibitors/2j-antennas",
    "https://www.worlddefenseshow.com/en/exhibitors/AC-00001",
    "https://www.idexuae.ae/exhibitors",
    "https://www.shieldafrica.com/en/exhibitors/",
    "https://www.milipol.com/en/exhibitors",
]

# ──────────────────────────────────────────────────────────────────────────────
# Excel styles
# ──────────────────────────────────────────────────────────────────────────────

NAVY   = PatternFill("solid", fgColor="1F3864")
BLUE2  = PatternFill("solid", fgColor="2E75B6")
GREEN2 = PatternFill("solid", fgColor="375623")
STRIPE = PatternFill("solid", fgColor="DCE6F1")
STRIPE2= PatternFill("solid", fgColor="E2EFDA")
HDR_FONT  = Font(bold=True, color="FFFFFF", size=11)
BODY_FONT = Font(size=10)
CENTER    = Alignment(horizontal="center", vertical="center", wrap_text=True)
TOP_WRAP  = Alignment(vertical="top", wrap_text=True)


def _write_header(ws, columns: list[tuple[str, int]], fill=NAVY) -> None:
    for col, (label, width) in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col, value=label)
        cell.fill      = fill
        cell.font      = HDR_FONT
        cell.alignment = CENTER
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"


# ──────────────────────────────────────────────────────────────────────────────
# SECTION 1 — TAVILY SEARCH
# ──────────────────────────────────────────────────────────────────────────────

def tavily_search(query: str, max_results: int = 10) -> dict:
    """
    POST https://api.tavily.com/search
    Returns {"results": [...], "elapsed": float, "error": str}
    """
    if not TAVILY_API_KEY:
        return {"results": [], "elapsed": 0.0, "error": "TAVILY_API_KEY not set"}

    url  = "https://api.tavily.com/search"
    body = {
        "api_key":      TAVILY_API_KEY,
        "query":        query,
        "max_results":  max_results,
        "search_depth": "basic",   # "basic" | "advanced"
        "include_answer":      False,
        "include_raw_content": False,
    }

    t0 = time.perf_counter()
    try:
        resp = httpx.post(url, json=body, timeout=30)
        elapsed = time.perf_counter() - t0

        if resp.status_code == 200:
            data = resp.json()
            return {
                "results": data.get("results", []),
                "elapsed": elapsed,
                "error":   "",
            }
        else:
            return {
                "results": [],
                "elapsed": elapsed,
                "error":   f"HTTP {resp.status_code}: {resp.text[:200]}",
            }
    except Exception as e:
        return {
            "results": [],
            "elapsed": time.perf_counter() - t0,
            "error":   f"{type(e).__name__}: {e}",
        }


def run_tavily_tests() -> list[dict]:
    """Run semua Tavily queries, return list of normalized records."""
    records: list[dict] = []
    seen_urls: set[str] = set()
    stats: list[dict]   = []

    console.print(Panel.fit(
        "[bold cyan]SECTION 1 — Tavily Search[/bold cyan]\n"
        f"[dim]{len(TAVILY_QUERIES)} queries  |  10 results each[/dim]",
        border_style="cyan",
    ))

    if not TAVILY_API_KEY:
        console.print("[bold red]TAVILY_API_KEY tidak ditemukan di .env — skip section ini.[/bold red]")
        return []

    with Progress(
        SpinnerColumn(),
        TextColumn("[progress.description]{task.description:<55}"),
        TimeElapsedColumn(),
        console=console,
    ) as progress:
        task = progress.add_task("Tavily search ...", total=len(TAVILY_QUERIES))

        for query in TAVILY_QUERIES:
            progress.update(task, description=f"[cyan]{query[:52]}[/cyan]")
            result = tavily_search(query)
            n_new  = 0

            for item in result["results"]:
                url   = item.get("url", "").strip()
                title = item.get("title", "").strip()
                if not url or url in seen_urls:
                    continue
                seen_urls.add(url)
                n_new += 1
                records.append({
                    "query":       query,
                    "title":       title,
                    "url":         url,
                    "domain":      urlparse(url).netloc.lstrip("www."),
                    "score":       round(item.get("score", 0.0), 4),
                    "snippet":     item.get("content", "")[:350],
                    "fetched_at":  datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    "elapsed_s":   round(result["elapsed"], 2),
                    "error":       result["error"],
                })

            stats.append({
                "query":   query,
                "results": n_new,
                "elapsed": round(result["elapsed"], 2),
                "error":   result["error"],
            })
            progress.advance(task)
            time.sleep(0.5)

    # Print mini summary
    table = Table(box=box.SIMPLE_HEAVY, header_style="bold white on dark_blue", show_lines=True)
    table.add_column("Query",   min_width=50, style="cyan")
    table.add_column("Results", justify="center", width=9)
    table.add_column("Time(s)", justify="center", width=9)
    table.add_column("Status",  justify="center", width=14)

    for s in stats:
        status = "[green]✅[/green]" if not s["error"] else f"[red]❌ {s['error'][:30]}[/red]"
        table.add_row(s["query"][:50], str(s["results"]), str(s["elapsed"]), status)

    console.print(table)
    console.print(f"  [bold]Total unique URLs dari Tavily:[/bold] [green]{len(records)}[/green]\n")
    return records


def write_tavily_sheet(wb: openpyxl.Workbook, records: list[dict]) -> None:
    ws = wb.create_sheet(title="Tavily Search")
    cols = [
        ("No",         5),
        ("Query",     52),
        ("Title",     50),
        ("URL",       60),
        ("Domain",    30),
        ("Score",     10),
        ("Snippet",   65),
        ("Time (s)",  10),
        ("Fetched At",20),
        ("Error",     30),
    ]
    _write_header(ws, cols, fill=BLUE2)

    for i, rec in enumerate(records, start=2):
        stripe = STRIPE if i % 2 == 0 else None
        values = [
            i - 1,
            rec["query"],
            rec["title"],
            rec["url"],
            rec["domain"],
            rec["score"],
            rec["snippet"],
            rec["elapsed_s"],
            rec["fetched_at"],
            rec["error"],
        ]
        for col, val in enumerate(values, start=1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.font      = BODY_FONT
            cell.alignment = TOP_WRAP
            if stripe:
                cell.fill = stripe
        ws.row_dimensions[i].height = 40


# ──────────────────────────────────────────────────────────────────────────────
# SECTION 2 — FIRECRAWL SCRAPE
# ──────────────────────────────────────────────────────────────────────────────

def firecrawl_scrape(url: str) -> dict:
    """
    POST https://api.firecrawl.dev/v1/scrape
    Returns {"url", "success", "markdown", "title", "description", "elapsed_s", "error", "metadata"}
    """
    if not FIRECRAWL_API_KEY:
        return {
            "url": url, "success": False, "markdown": "",
            "title": "", "description": "", "elapsed_s": 0.0,
            "error": "FIRECRAWL_API_KEY not set", "metadata": {},
        }

    endpoint = "https://api.firecrawl.dev/v1/scrape"
    headers  = {
        "Authorization": f"Bearer {FIRECRAWL_API_KEY}",
        "Content-Type":  "application/json",
    }
    body = {
        "url":     url,
        "formats": ["markdown", "extract"],
        "extract": {
            "schema": {
                "type": "object",
                "properties": {
                    "company_name":  {"type": "string"},
                    "website":       {"type": "string"},
                    "email":         {"type": "string"},
                    "phone":         {"type": "string"},
                    "country":       {"type": "string"},
                    "description":   {"type": "string"},
                    "booth_number":  {"type": "string"},
                },
            }
        },
    }

    t0 = time.perf_counter()
    try:
        resp    = httpx.post(endpoint, headers=headers, json=body, timeout=60)
        elapsed = time.perf_counter() - t0

        if resp.status_code == 200:
            data     = resp.json()
            inner    = data.get("data", {})
            meta     = inner.get("metadata", {})
            markdown = inner.get("markdown", "")
            extract  = inner.get("extract", {}) or {}
            return {
                "url":         url,
                "success":     data.get("success", False),
                "markdown":    markdown,
                "markdown_len": len(markdown),
                "title":       meta.get("title", ""),
                "description": meta.get("description", ""),
                "elapsed_s":   round(elapsed, 2),
                "error":       "",
                "extract":     extract,
                "metadata":    meta,
            }
        else:
            return {
                "url": url, "success": False, "markdown": "",
                "markdown_len": 0, "title": "", "description": "",
                "elapsed_s": round(elapsed, 2),
                "error": f"HTTP {resp.status_code}: {resp.text[:300]}",
                "extract": {}, "metadata": {},
            }
    except Exception as e:
        return {
            "url": url, "success": False, "markdown": "",
            "markdown_len": 0, "title": "", "description": "",
            "elapsed_s": round(time.perf_counter() - t0, 2),
            "error": f"{type(e).__name__}: {e}",
            "extract": {}, "metadata": {},
        }


def run_firecrawl_tests() -> list[dict]:
    """Scrape semua test URLs dengan Firecrawl."""
    records: list[dict] = []

    console.print(Panel.fit(
        "[bold green]SECTION 2 — Firecrawl Scrape[/bold green]\n"
        f"[dim]{len(FIRECRAWL_TEST_URLS)} URLs[/dim]",
        border_style="green",
    ))

    if not FIRECRAWL_API_KEY:
        console.print("[bold red]FIRECRAWL_API_KEY tidak ditemukan di .env — skip section ini.[/bold red]")
        return []

    with Progress(
        SpinnerColumn(),
        TextColumn("[progress.description]{task.description:<60}"),
        TimeElapsedColumn(),
        console=console,
    ) as progress:
        task = progress.add_task("Firecrawl scrape ...", total=len(FIRECRAWL_TEST_URLS))

        for url in FIRECRAWL_TEST_URLS:
            progress.update(task, description=f"[green]{url[:58]}[/green]")
            rec = firecrawl_scrape(url)
            records.append(rec)
            progress.advance(task)
            time.sleep(1.0)  # rate limit

    # Mini summary table
    table = Table(box=box.SIMPLE_HEAVY, header_style="bold white on dark_green", show_lines=True)
    table.add_column("URL",      min_width=55, style="green")
    table.add_column("Success",  justify="center", width=9)
    table.add_column("MD Len",   justify="center", width=9)
    table.add_column("Time(s)",  justify="center", width=9)
    table.add_column("Status",   justify="center", width=18)

    for rec in records:
        ok     = rec["success"]
        status = "[green]✅[/green]" if ok else f"[red]❌ {rec['error'][:30]}[/red]"
        table.add_row(
            rec["url"][:55],
            "✅" if ok else "❌",
            str(rec.get("markdown_len", 0)),
            str(rec["elapsed_s"]),
            status,
        )

    console.print(table)

    ok_count = sum(1 for r in records if r["success"])
    console.print(f"  [bold]Sukses:[/bold] [green]{ok_count}/{len(records)}[/green]\n")

    return records


def write_firecrawl_sheet(wb: openpyxl.Workbook, records: list[dict]) -> None:
    ws = wb.create_sheet(title="Firecrawl Scrape")
    cols = [
        ("No",             5),
        ("URL",           60),
        ("Success",       10),
        ("Title",         45),
        ("Description",   50),
        ("Markdown Len",  13),
        ("company_name",  30),
        ("website",       30),
        ("email",         25),
        ("phone",         20),
        ("country",       15),
        ("booth_number",  15),
        ("Time (s)",       9),
        ("Error",         35),
        ("Markdown (500)", 80),
    ]
    _write_header(ws, cols, fill=GREEN2)

    for i, rec in enumerate(records, start=2):
        stripe  = STRIPE2 if i % 2 == 0 else None
        extract = rec.get("extract", {}) or {}
        values  = [
            i - 1,
            rec["url"],
            "YES" if rec["success"] else "NO",
            rec.get("title", ""),
            rec.get("description", ""),
            rec.get("markdown_len", 0),
            extract.get("company_name", ""),
            extract.get("website", ""),
            extract.get("email", ""),
            extract.get("phone", ""),
            extract.get("country", ""),
            extract.get("booth_number", ""),
            rec["elapsed_s"],
            rec.get("error", ""),
            rec.get("markdown", "")[:500],
        ]
        for col, val in enumerate(values, start=1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.font      = BODY_FONT
            cell.alignment = TOP_WRAP
            if stripe:
                cell.fill = stripe
        ws.row_dimensions[i].height = 60


def write_firecrawl_markdown_sheet(wb: openpyxl.Workbook, records: list[dict]) -> None:
    """Sheet tambahan: full markdown per URL (1 row = 1 URL)."""
    ws = wb.create_sheet(title="Firecrawl Markdown")
    cols = [
        ("No",     5),
        ("URL",   60),
        ("Chars", 10),
        ("Full Markdown", 120),
    ]
    _write_header(ws, cols, fill=GREEN2)

    for i, rec in enumerate(records, start=2):
        md     = rec.get("markdown", "")
        values = [i - 1, rec["url"], len(md), md[:8000]]
        for col, val in enumerate(values, start=1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.font      = BODY_FONT
            cell.alignment = TOP_WRAP
        ws.row_dimensions[i].height = 120


# ──────────────────────────────────────────────────────────────────────────────
# SECTION 3 — COMBINED SUMMARY SHEET
# ──────────────────────────────────────────────────────────────────────────────

def write_summary_sheet(
    wb: openpyxl.Workbook,
    tavily_records: list[dict],
    firecrawl_records: list[dict],
) -> None:
    ws = wb.create_sheet(title="Summary", index=0)

    ws.merge_cells("A1:E1")
    c = ws["A1"]
    c.value     = f"Tavily + Firecrawl Test — {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    c.font      = Font(bold=True, size=14, color="FFFFFF")
    c.fill      = NAVY
    c.alignment = CENTER
    ws.row_dimensions[1].height = 30

    # Key status
    rows = [
        ("TAVILY_API_KEY",    "✅ Set" if TAVILY_API_KEY    else "❌ Missing"),
        ("FIRECRAWL_API_KEY", "✅ Set" if FIRECRAWL_API_KEY else "❌ Missing"),
        ("", ""),
        ("Tavily unique URLs",            str(len(tavily_records))),
        ("Firecrawl success",             str(sum(1 for r in firecrawl_records if r.get("success")))),
        ("Firecrawl failed",              str(sum(1 for r in firecrawl_records if not r.get("success")))),
        ("Avg markdown len (success)",    str(
            round(sum(r.get("markdown_len", 0) for r in firecrawl_records if r.get("success")) /
                  max(1, sum(1 for r in firecrawl_records if r.get("success"))))
        ) if firecrawl_records else "0"),
    ]

    for row_idx, (k, v) in enumerate(rows, start=3):
        ws.cell(row=row_idx, column=1, value=k).font = Font(bold=True, size=11)
        ws.cell(row=row_idx, column=2, value=v).font = Font(size=11)

    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 25


# ──────────────────────────────────────────────────────────────────────────────
# EXPORT EXCEL
# ──────────────────────────────────────────────────────────────────────────────

def export_excel(tavily_records: list[dict], firecrawl_records: list[dict]) -> str:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    ts       = datetime.now().strftime("%Y%m%d_%H%M%S")
    filepath = OUTPUT_DIR / f"test_firecrawl_tavily_{ts}.xlsx"

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    write_summary_sheet(wb, tavily_records, firecrawl_records)

    if tavily_records:
        write_tavily_sheet(wb, tavily_records)

    if firecrawl_records:
        write_firecrawl_sheet(wb, firecrawl_records)
        write_firecrawl_markdown_sheet(wb, firecrawl_records)

    wb.save(str(filepath))
    return str(filepath.resolve())


# ──────────────────────────────────────────────────────────────────────────────
# MAIN
# ──────────────────────────────────────────────────────────────────────────────

def main() -> int:
    console.print(Panel.fit(
        "[bold white]Tavily Search + Firecrawl Scrape — Test Script[/bold white]\n"
        "[dim]Standalone · baca key dari .env · export ke Excel[/dim]",
        border_style="white",
        padding=(0, 3),
    ))

    # Show key status upfront
    console.print(f"  TAVILY_API_KEY    : {'[green]✅ Set[/green]' if TAVILY_API_KEY    else '[red]❌ Missing (set di .env)[/red]'}")
    console.print(f"  FIRECRAWL_API_KEY : {'[green]✅ Set[/green]' if FIRECRAWL_API_KEY else '[red]❌ Missing (set di .env)[/red]'}")
    console.print()

    if not TAVILY_API_KEY and not FIRECRAWL_API_KEY:
        console.print(
            "[bold red]Kedua API key tidak ditemukan.[/bold red]\n"
            "Tambahkan ke .env:\n"
            "  [yellow]TAVILY_API_KEY=tvly-xxxx[/yellow]\n"
            "  [yellow]FIRECRAWL_API_KEY=fc-xxxx[/yellow]"
        )
        return 1

    tavily_records    = run_tavily_tests()
    firecrawl_records = run_firecrawl_tests()

    console.print("[dim]Exporting to Excel ...[/dim]")
    excel_path = export_excel(tavily_records, firecrawl_records)

    console.print(Panel.fit(
        f"[bold]Selesai![/bold]\n\n"
        f"  Tavily URLs    : [green]{len(tavily_records)}[/green]\n"
        f"  Firecrawl OK   : [green]{sum(1 for r in firecrawl_records if r.get('success'))}[/green] / {len(firecrawl_records)}\n"
        f"  Excel output   : [cyan]{excel_path}[/cyan]",
        border_style="green",
        padding=(0, 3),
    ))
    return 0


if __name__ == "__main__":
    sys.exit(main())
